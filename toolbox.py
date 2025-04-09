from fuzzylogic.classes import Domain
from fuzzylogic.functions import R, S, trapezoid
import socket
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

PRESET_Value = 0xFFFF
POLYNOMIAL = 0x8408

def pointArrangement(minVal, maxVal):
    points = [0] * 8
    points[0] = minVal
    points[1] = maxVal / 5
    points[2] = points[1] + minVal
    points[3] = points[1] * 2
    points[4] = points[3] + minVal
    points[5] = points[1] * 3
    points[6] = points[5] + minVal
    points[7] = points[1] * 4
    return points

def generate_domain(name, minVal, maxVal, labels, res=0.1):
    points = pointArrangement(minVal, maxVal)
    domain = Domain(name, 0, maxVal, res=res)
    setattr(domain, labels[0], S(points[0], points[1]))
    setattr(domain, labels[1], trapezoid(points[0], points[1], points[2], points[3], c_m=1))
    setattr(domain, labels[2], trapezoid(points[2], points[3], points[4], points[5], c_m=1))
    setattr(domain, labels[3], trapezoid(points[4], points[5], points[6], points[7], c_m=1))
    setattr(domain, labels[4], R(points[6], points[7]))
    return domain

def crc(cmd):
    cmd = bytes.fromhex(cmd)
    viCrcValue = PRESET_Value
    for x in cmd:
        viCrcValue ^= x
        for _ in range(8):
            if viCrcValue & 0x0001:
                viCrcValue = (viCrcValue >> 1) ^ POLYNOMIAL
            else:
                viCrcValue >>= 1
    crc_H = (viCrcValue >> 8) & 0xFF
    crc_L = viCrcValue & 0xFF
    return cmd + bytes([crc_L, crc_H])

def send_cmd(cmd, host='192.168.1.192', port=6000):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(2)
            s.connect((host, port))
            message = crc(cmd)
            s.sendall(message)
            data = s.recv(64)
            if not data:
                return ""
            response_hex = data.hex().upper()
            if len(response_hex) % 2 == 0:
                return [response_hex[i*2:i*2+2] for i in range(len(response_hex)//2)]
            else:
                return ['FF']
    except (socket.timeout, socket.error):
        return "ERROR"

def read_power_meter(client, ct_ratio=150):
    try:
        result = client.read_holding_registers(address=0x0000, count=10, slave=2)
        if not result.isError():
            regs = result.registers
            v_ab = regs[0] / 10.0
            v_bc = regs[1] / 10.0
            v_ca = regs[2] / 10.0
            i_a = regs[3] / 100.0 * ct_ratio
            i_b = regs[4] / 100.0 * ct_ratio
            i_c = regs[5] / 100.0 * ct_ratio
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            return (timestamp, v_ab, v_bc, v_ca, i_a, i_b, i_c)
        else:
            return None
    except:
        return None

def log_power_to_excel(client, ct_ratio=150):
    filename = f"power_log_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    sheet_name = "PowerLog"

    data = read_power_meter(client, ct_ratio)
    if data is None:
        return

    try:
        wb = openpyxl.load_workbook(filename)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Timestamp", "V_AB", "V_BC", "V_CA", "I_A", "I_B", "I_C"])
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(["Timestamp", "V_AB", "V_BC", "V_CA", "I_A", "I_B", "I_C"])

    ws.append(data)
    wb.save(filename)