import time, socket, json, os, random
import pandas as pd
from datetime import datetime
from toolbox import generate_domain
from fuzzylogic.classes import Rule
from pymodbus.client import ModbusSerialClient
from openpyxl import Workbook, load_workbook

# Global random mode (True for testing)
randomify = True

defaultDur = 3000
client = ModbusSerialClient(port='/dev/ttyACM0', baudrate=9600)
host, port_num = '192.168.1.192', 6000
INVENTORY = '04 FF 0F'
PRESET_Value, POLYNOMIAL = 0xFFFF, 0x8408
timeout = 30
listNull = ['05', '00', '0F', 'FB', 'E2', 'A7']

# Membership Function
durasiOven = generate_domain("Durasi Oven (Menit)", 120, 1440, ["sangat_sebentar", "sebentar", "sedang", "lama", "sangat_lama"])
pastLub = generate_domain("Rentang Lubrikasi Terakhir (Jam)", 12, 672, ["sangat_sebentar", "sebentar", "sedang", "lama", "sangat_lama"])
lastLub = generate_domain("Durasi Lubrikasi Terakhir (Milisecond)", 50, 6000, ["tidak_spray", "sebentar", "sedang", "lama", "sangat_lama"])
lubDur = generate_domain("Durasi Lubrikasi (Milisecond)", 20, 6000, ["tidak_spray", "sebentar", "sedang", "lama", "sangat_lama"])

# Fuzzy Rule
df = pd.read_csv("rulebases.csv")
rule_dict = {(row["Durasi_Oven"], row["Lubrikasi_Terakhir"], row["Durasi_Terakhir"]): row["Durasi_Lubrikasi"] for _, row in df.iterrows()}
rules = [Rule({(getattr(durasiOven, do), getattr(pastLub, lt), getattr(lastLub, dt)): getattr(lubDur, dl)})
         for (do, lt, dt), dl in rule_dict.items()]
final_rule = sum(rules)

def crc(cmd):
    cmd = bytes.fromhex(cmd)
    viCrcValue = PRESET_Value
    for x in cmd:
        viCrcValue ^= x
        for _ in range(8):
            viCrcValue = (viCrcValue >> 1) ^ POLYNOMIAL if viCrcValue & 0x0001 else viCrcValue >> 1
    return cmd + bytes([viCrcValue & 0xFF, (viCrcValue >> 8) & 0xFF])

def send_cmd(cmd):
    try:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(2)
            s.connect((host, port_num))
            s.sendall(crc(cmd))
            data = s.recv(64)
            if not data: return ""
            return [data.hex().upper()[i*2:(i+1)*2] for i in range(len(data)//2)]
    except:
        return "ERROR"

def scan_rfid():
    start_time = time.time()
    last_state = ""
    while True:
        result = send_cmd(INVENTORY)
        if result != last_state:
            if result == listNull:
                print("Menunggu tag...")
            elif result == "ERROR":
                print("Koneksi RFID error.")
            else:
                return int(result[6] + result[7], 16)
            last_state = result
        if time.time() - start_time > timeout:
            print("Timeout RFID")
            return 17
        time.sleep(0.04)

def log_to_excel(cart_data, voltage, current):
    filename = datetime.now().strftime("log_%Y-%m-%d.xlsx")
    if not os.path.exists(filename):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Cart Logs"
        ws1.append(["Timestamp", "CartID", "lastTS", "lastLubTS", "lastLubDur"])
        ws2 = wb.create_sheet("Power Logs")
        ws2.append(["Timestamp", "Voltage", "Current"])
        wb.save(filename)

    wb = load_workbook(filename)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    wb["Cart Logs"].append([ts, cart_data["cartID"], cart_data["lastTS"], cart_data["lastLubTS"], cart_data["lastLubDur"]])
    wb["Power Logs"].append([ts, voltage, current])
    wb.save(filename)

def calc_lub(idTag):
    strJsonName = f"cart{idTag}.json"
    with open(strJsonName, "r") as file:
        data = json.load(file)

    data["lastTS"] = datetime.fromisoformat(data["lastLubTS"])
    lastScan = round((datetime.now() - data["lastTS"]).total_seconds() / 60, 3)
    lastLubSpan = round((datetime.now() - datetime.fromisoformat(data["lastLubTS"])).total_seconds() / 3600, 3)
    lastLubDur = data["lastLubDur"]

    values = {durasiOven: lastScan, pastLub: lastLubSpan, lastLub: lastLubDur}
    result = round(final_rule(values))

    if result < 500:
        result = 0
        data["lastLubDur"] = 0
    else:
        data["lastLubDur"] = result
    
    if idTag == 17:
        result = defaultDur
        data["lastLubDur"] = defaultDur

    data["lastTS"] = datetime.now().isoformat()
    data["lastLubTS"] = datetime.now().isoformat()

    with open(strJsonName, "w") as file:
        json.dump(data, file, indent=4)

    data["cartID"] = idTag
    return result, data

def read_power_meter():
    if randomify:
        return round(random.uniform(210, 240), 2), round(random.uniform(2.0, 4.0), 2)
    try:
        voltage = client.read_holding_registers(0x0000, 2, slave=2).registers
        current = client.read_holding_registers(0x0006, 2, slave=2).registers
        voltage_val = (voltage[0] << 16 | voltage[1]) / 10
        current_val = (current[0] << 16 | current[1]) / 1000
        return round(voltage_val, 2), round(current_val, 3)
    except:
        return 0, 0

def main():
    prev_stat = 0
    while True:
        try:
            stat = client.read_holding_registers(502, 1, slave=1).registers[0]
        except:
            stat = 0

        voltage, current = read_power_meter()

        if prev_stat == 0 and stat == 1:
            print("Kereta terdeteksi")
            try:
                idTag = scan_rfid()
            except:
                idTag = 17

            dur, cart_data = calc_lub(idTag)
            print(f"Lubrikasi: {dur} ms")
            client.write_register(address=501, value=dur, slave=1)
            log_to_excel(cart_data, voltage, current)

        else:
            print("Menunggu kereta...")

        prev_stat = stat
        time.sleep(1)

if __name__ == "__main__":
    main()